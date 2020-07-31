from openpyxl import load_workbook
import sys
import codecs

sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

workbook = load_workbook('./kfc_ph_link.xlsx')  #找到需要xlsx文件的位置
booksheet = workbook.active

#获取sheet页的行数据
rows = booksheet.rows
#获取sheet页的列数据
columns = booksheet.columns

i = 0
list = []
for row in rows:
  i = i + 1
  name = booksheet.cell(row=i, column=3).value
  value = booksheet.cell(row=i, column=2).value
  if name and value:
    list.append([name, value])


# 迭代所有的行
with open('kfc_ph_link.json', 'w') as file:
  file.write('{' + '\n')
  for i, item in enumerate(list):
    file.write('  "{}": "{}"{}\n'.format(item[0], item[1], ',' if i < len(list)-1 else ''))

  file.write('}')
  
